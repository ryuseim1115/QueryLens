from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

wb = Workbook()

# ── スタイル定数 ─────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
SUBHEAD_FILL = PatternFill("solid", fgColor="BDD7EE")
ALT_FILL = PatternFill("solid", fgColor="DEEAF1")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

WHITE_BOLD = Font(name="Meiryo UI", bold=True, color="FFFFFF", size=10)
DARK_BOLD = Font(name="Meiryo UI", bold=True, color="1F4E79", size=10)
NORMAL_FONT = Font(name="Meiryo UI", size=10)

THIN = Side(style="thin", color="9DC3E6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(cell, title):
    cell.value = title
    cell.font = WHITE_BOLD
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER


def style_section(cell, title):
    cell.value = title
    cell.font = WHITE_BOLD
    cell.fill = SECTION_FILL
    cell.alignment = LEFT
    cell.border = BORDER


def style_subhead(cell, title):
    cell.value = title
    cell.font = DARK_BOLD
    cell.fill = SUBHEAD_FILL
    cell.alignment = CENTER
    cell.border = BORDER


def style_cell(cell, value, alt=False):
    cell.value = value
    cell.font = NORMAL_FONT
    cell.fill = ALT_FILL if alt else WHITE_FILL
    cell.alignment = LEFT
    cell.border = BORDER


def set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def freeze(ws, cell="A2"):
    ws.freeze_panes = cell


# ════════════════════════════════════════════════════════════
# Sheet 1: 概要
# ════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "概要"
ws1.row_dimensions[1].height = 30

style_header(ws1["B2"], "QueryLens 詳細設計書 — Post-Analyzer")
ws1.merge_cells("B2:G2")
ws1["B2"].font = Font(name="Meiryo UI", bold=True, color="FFFFFF", size=14)

rows = [
    ("ドキュメント名", "QueryLens_Detailed_Design_Post_Analyzer"),
    ("対象ブランチ", "refactor-subquery-analysis / feat-display-query-result"),
    ("作成日", "2026-05-02"),
    (
        "概要",
        "SQLクエリを受け取り、サブクエリ単位に分解・実行し結果を表示する\n"
        "Post-Analyzerコンポーネントの詳細設計。",
    ),
]
for i, (k, v) in enumerate(rows, start=4):
    alt = i % 2 == 0
    style_subhead(ws1.cell(i, 2), k)
    style_cell(ws1.cell(i, 3), v, alt)
    ws1.merge_cells(f"C{i}:G{i}")
    ws1.row_dimensions[i].height = 36

set_col_widths(ws1, {"B": 22, "C": 60, "D": 10, "E": 10, "F": 10, "G": 10})


# ════════════════════════════════════════════════════════════
# Sheet 2: スキーマ定義
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("スキーマ定義")
ws2.row_dimensions[1].height = 30

headers = ["クラス名", "フィールド名", "型", "必須", "説明"]
for ci, h in enumerate(headers, start=1):
    style_subhead(ws2.cell(1, ci), h)

schemas = [
    # (クラス名, フィールド, 型, 必須, 説明)
    ("QueryInfo", "database_type", "str", "○", "データベース種別"),
    ("QueryInfo", "query", "str", "○", "実行するSQLクエリ"),
    ("TableInfo", "name", "str | None", "×", "テーブル名"),
    ("TableInfo", "alias", "str | None", "×", "テーブルエイリアス"),
    ("SubqueryAnalyzeResult", "start_index", "int", "○", "クエリ文字列内の開始位置"),
    ("SubqueryAnalyzeResult", "end_index", "int", "○", "クエリ文字列内の終了位置"),
    ("SubqueryAnalyzeResult", "query", "str", "○", "サブクエリのSQL文"),
    ("SubqueryAnalyzeResult", "depth", "int", "○", "ネスト深さ（0=最外）"),
    (
        "SubqueryAnalyzeResult",
        "tables_name_alias",
        "list[TableInfo]",
        "○",
        "FROM/JOINで使用するテーブル一覧",
    ),
    (
        "SubqueryAnalyzeResult",
        "parent_alias",
        "str | None",
        "×",
        "親クエリ内でのエイリアス名",
    ),
    (
        "SubqueryAnalyzeResult",
        "result",
        "list[dict[str, Any]]",
        "○",
        "クエリ実行結果（初期値=[]）",
    ),
    (
        "RunQueryResponse",
        "subqueries",
        "list[SubqueryAnalyzeResult]",
        "○",
        "分析・実行済みサブクエリ一覧",
    ),
]

for ri, row in enumerate(schemas, start=2):
    alt = ri % 2 == 0
    for ci, val in enumerate(row, start=1):
        style_cell(ws2.cell(ri, ci), val, alt)

set_col_widths(ws2, {"A": 24, "B": 22, "C": 26, "D": 8, "E": 40})
freeze(ws2)


# ════════════════════════════════════════════════════════════
# Sheet 3: バックエンド クラス・関数
# ════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("バックエンド")
ws3.row_dimensions[1].height = 30

headers = ["モジュール", "クラス / 関数", "種別", "引数", "戻り値", "処理概要", "依存"]
for ci, h in enumerate(headers, start=1):
    style_subhead(ws3.cell(1, ci), h)

backend_rows = [
    # routers
    (
        "routers/run_query.py",
        "run_query",
        "関数",
        "body: QueryInfo",
        "RunQueryResponse",
        "バリデーション→構造解析→深さソート→クエリ実行の一連フローを実行。"
        "エラー時はHTTP 400を返す。",
        "QueryValidator, QueryStructureAnalyzer, SortSubqueryByDepthDesc,"
        " SubqueryRunner",
    ),
    # validators
    (
        "validators/query_validator.py",
        "QueryValidator",
        "クラス",
        "database_type: str, query: str",
        "—",
        "SQLの構文チェックとテーブル存在確認を行うバリデータ。",
        "sqlglot, get_connection",
    ),
    (
        "validators/query_validator.py",
        "  validate()",
        "メソッド",
        "—",
        "sqlglot.Expression",
        "SELECT文であることを確認し、参照テーブルがDBに存在するかを検証。",
        "sqlglot",
    ),
    (
        "validators/query_validator.py",
        "  _validate_tables()",
        "メソッド",
        "expression: sqlglot.Expression",
        "None",
        "CTE名を除いた実テーブル名の存在をDuckDBのSHOW TABLESで検証。",
        "get_connection",
    ),
    # query_structure_analyzer
    (
        "services/.../query_structure_analyzer.py",
        "QueryStructureAnalyzer",
        "クラス",
        "query: str",
        "—",
        "クエリを分解しSubqueryAnalyzeResultリストに変換するファサード。",
        "AnalyzeSubquery, TableInfo",
    ),
    (
        "services/.../query_structure_analyzer.py",
        "  execute()",
        "メソッド",
        "—",
        "SubqueryAnalyzeResultList",
        "AnalyzeSubqueryの結果をSubqueryAnalyzeResultスキーマへマッピング。",
        "AnalyzeSubquery",
    ),
    # subquery_builder
    (
        "services/.../subquery_builder.py",
        "Subquery",
        "データクラス",
        "—",
        "—",
        "サブクエリの位置・テキスト・深さ・テーブル情報・親エイリアスを保持。",
        "—",
    ),
    (
        "services/.../subquery_builder.py",
        "AnalyzeSubquery",
        "クラス",
        "query: str",
        "—",
        "SQLをトークン化し、サブクエリ範囲・深さ・テーブルを解析するコアクラス。",
        "find_subquery_ranges, get_subquery_depths, extract_tables_with_alias",
    ),
    (
        "services/.../subquery_builder.py",
        "  execute()",
        "メソッド",
        "—",
        "list[Subquery]",
        "_build()を呼び出してSubqueryリストを返す。",
        "—",
    ),
    # subquery_table_extractor
    (
        "services/.../subquery_table_extractor.py",
        "extract_tables_with_alias()",
        "関数",
        "query: str",
        "list[tuple[str|None, str|None]]",
        "FROM/JOINからテーブル名とエイリアスのペアを抽出。",
        "sqlglot",
    ),
    # subquery_depth_analyzer
    (
        "services/.../subquery_depth_analyzer.py",
        "get_subquery_depths()",
        "関数",
        "ranges: list[tuple[int,int]]",
        "list[int]",
        "各範囲を包含する上位範囲の数をカウントしネスト深さを算出。",
        "—",
    ),
    # subquery_range_finder
    (
        "services/.../subquery_range_finder.py",
        "find_subquery_ranges()",
        "関数",
        "query: str, tokens: list",
        "dict[tuple[int,int], str|None]",
        "トークンをスキャンしサブクエリの(開始,終了)→エイリアスのマッピングを返す。",
        "sqlglot",
    ),
    (
        "services/.../subquery_range_finder.py",
        "find_cte_ranges()",
        "関数",
        "tokens: list",
        "dict[tuple[int,int], str]",
        "AS SELECT パターンを検出し CTE の範囲→CTE名のマッピングを返す。",
        "sqlglot",
    ),
    # sort_subquery
    (
        "services/.../sort_subquery.py",
        "SortSubqueryByDepthDesc",
        "クラス",
        "subqueries: SubqueryAnalyzeResultList",
        "—",
        "サブクエリを深さ降順にソートするクラス。深いものから実行するために使用。",
        "—",
    ),
    (
        "services/.../sort_subquery.py",
        "  execute()",
        "メソッド",
        "—",
        "SubqueryAnalyzeResultList",
        "depth降順でソートしたリストを返す。",
        "—",
    ),
    # db
    (
        "db/create_csv_tables.py",
        "create_csv_tables()",
        "関数",
        "csv_files_dir: str, csv_files: list[str]",
        "None",
        "CSVファイルをDuckDBテーブルとして登録（CREATE OR REPLACE TABLE）。",
        "get_connection",
    ),
    (
        "db/run_subqueries.py",
        "run_subqueries()",
        "関数",
        "subqueries: SubqueryAnalyzeResultList",
        "SubqueryAnalyzeResultList",
        "各サブクエリをDuckDBで実行し結果を辞書リストで格納。失敗時はValueErrorを送出。",
        "get_connection",
    ),
]

for ri, row in enumerate(backend_rows, start=2):
    alt = ri % 2 == 0
    for ci, val in enumerate(row, start=1):
        style_cell(ws3.cell(ri, ci), val, alt)
    ws3.row_dimensions[ri].height = 40

set_col_widths(ws3, {"A": 36, "B": 32, "C": 12, "D": 36, "E": 36, "F": 56, "G": 44})
freeze(ws3)


# ════════════════════════════════════════════════════════════
# Sheet 4: フロントエンド 関数
# ════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("フロントエンド")
ws4.row_dimensions[1].height = 30

headers = ["ファイル", "関数名", "引数", "戻り値", "処理概要", "依存"]
for ci, h in enumerate(headers, start=1):
    style_subhead(ws4.cell(1, ci), h)

frontend_rows = [
    (
        "pages/Result.js",
        "（エントリーポイント）",
        "—",
        "—",
        "sessionStorageからクエリ・サブクエリを取得し、表示関数を順次実行。",
        "DisplayTables, DisplayQuery, DisplayLines, DisplayQueryResult",
    ),
    (
        "display/DisplayQuery.js",
        "displayQuery()",
        "query: string",
        "void",
        "'query-display'要素にクエリテキストを表示。",
        "—",
    ),
    (
        "display/DisplayQuery.js",
        "highlightQuery()",
        "start_index: number, end_index: number",
        "void",
        "クエリの指定範囲をmark要素でハイライト表示。",
        "—",
    ),
    (
        "display/DisplayQuery.js",
        "escapeHtml()",
        "text: string",
        "string",
        "HTML特殊文字（&, <, >）をエスケープ。",
        "—",
    ),
    (
        "display/DisplayTables.js",
        "displayTables()",
        "subqueries: SubqueryAnalyzeResult[]",
        "void",
        "サブクエリを深さごとにグループ化し、テーブルカード一覧をDOMに追加。クリックでクエリハイライト。",
        "highlightQuery",
    ),
    (
        "display/DisplayTables.js",
        "createDepthGroup()",
        "depth: number",
        "HTMLElement",
        "指定深さの深さグループdiv要素を生成。",
        "—",
    ),
    (
        "display/DisplayLines.js",
        "displayLines()",
        "subqueries: SubqueryAnalyzeResult[]",
        "void",
        "canvas要素を生成し、各サブクエリから親エイリアスへ結線を描画。",
        "findParentAliasEl",
    ),
    (
        "display/DisplayQueryResult.js",
        "displayQueryResult()",
        "subqueries: SubqueryAnalyzeResult[]",
        "void",
        "各サブクエリの親エイリアス要素にクリックイベントを登録。",
        "findParentAliasEl, renderQueryResult",
    ),
    (
        "display/DisplayQueryResult.js",
        "renderQueryResult()",
        "subquery: SubqueryAnalyzeResult",
        "void",
        "クリック時に実行結果テーブルを再描画。結果が空の場合はnoResult()を呼ぶ。",
        "buildHeaderRow, buildDataRow, noResult",
    ),
    (
        "display/DisplayQueryResult.js",
        "buildHeaderRow()",
        "columns: string[]",
        "HTMLElement",
        "カラム名のth要素を持つtr要素を生成。",
        "—",
    ),
    (
        "display/DisplayQueryResult.js",
        "buildDataRow()",
        "row: object, columns: string[]",
        "HTMLElement",
        "行データのtd要素を持つtr要素を生成。",
        "—",
    ),
    (
        "display/DisplayQueryResult.js",
        "noResult()",
        "—",
        "void",
        "'no-result'要素を表示（0件メッセージ）。",
        "—",
    ),
    (
        "utility.js",
        "findParentAliasEl()",
        "subquery: SubqueryAnalyzeResult",
        "HTMLElement | null",
        "親の深さグループからdata-aliasが一致するテーブル要素を検索して返す。",
        "—",
    ),
]

for ri, row in enumerate(frontend_rows, start=2):
    alt = ri % 2 == 0
    for ci, val in enumerate(row, start=1):
        style_cell(ws4.cell(ri, ci), val, alt)
    ws4.row_dimensions[ri].height = 40

set_col_widths(ws4, {"A": 34, "B": 26, "C": 36, "D": 18, "E": 60, "F": 44})
freeze(ws4)


# ════════════════════════════════════════════════════════════
# Sheet 5: 処理フロー
# ════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("処理フロー")

style_header(ws5["B2"], "APIリクエスト処理フロー")
ws5.merge_cells("B2:E2")
ws5["B2"].font = Font(name="Meiryo UI", bold=True, color="FFFFFF", size=12)

flow = [
    ("1", "クライアント", "POST /run-query", "QueryInfo（database_type, query）を送信"),
    ("2", "QueryValidator", "validate()", "SQL構文チェック・テーブル存在確認"),
    ("3", "QueryStructureAnalyzer", "execute()", "クエリをサブクエリ単位に分解"),
    ("4", "AnalyzeSubquery", "_build()", "範囲抽出・深さ計算・テーブル抽出"),
    (
        "5",
        "SortSubqueryByDepthDesc",
        "execute()",
        "depth降順ソート（深いものから実行）",
    ),
    ("6", "SubqueryRunner", "execute()", "run_subqueries()でDuckDB実行"),
    (
        "7",
        "APIレスポンス",
        "RunQueryResponse",
        "SubqueryAnalyzeResultList をJSONで返却",
    ),
    ("8", "Result.js", "displayQuery()", "クエリ全文を表示"),
    ("9", "Result.js", "displayTables()", "テーブルカード一覧を描画"),
    ("10", "Result.js", "displayLines()", "サブクエリ間の依存線をcanvasに描画"),
    ("11", "Result.js", "displayQueryResult()", "クリックで実行結果を表示（0件対応）"),
]

step_headers = ["ステップ", "コンポーネント", "呼び出し", "内容"]
for ci, h in enumerate(step_headers, start=2):
    style_subhead(ws5.cell(4, ci), h)

for ri, row in enumerate(flow, start=5):
    alt = ri % 2 == 0
    for ci, val in enumerate(row, start=2):
        style_cell(ws5.cell(ri, ci), val, alt)
    ws5.row_dimensions[ri].height = 32

set_col_widths(ws5, {"B": 10, "C": 30, "D": 28, "E": 60})
freeze(ws5, "B5")


# ════════════════════════════════════════════════════════════
# 保存
# ════════════════════════════════════════════════════════════
out_path = "QueryLens_Detailed_Design_Post_Analyzer.xlsx"
wb.save(out_path)
print(f"生成完了: {out_path}")
